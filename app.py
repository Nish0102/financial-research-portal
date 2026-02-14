import os
import json
import re
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import tempfile

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024  # 25MB max
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()

ALLOWED_EXTENSIONS = {'pdf', 'txt', 'docx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_pdf(file_path):
    """Extract text from PDF using pdfplumber"""
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
    except Exception as e:
        return None, f"Error reading PDF: {str(e)}"
    return text, None

def extract_text_from_file(file_path):
    """Extract text from uploaded file"""
    file_ext = file_path.rsplit('.', 1)[1].lower()
    
    if file_ext == 'pdf':
        return extract_text_from_pdf(file_path)
    elif file_ext == 'txt':
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read(), None
        except Exception as e:
            return None, f"Error reading text file: {str(e)}"
    elif file_ext == 'docx':
        try:
            from docx import Document
            doc = Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
            return text, None
        except Exception as e:
            return None, f"Error reading DOCX: {str(e)}"
    
    return None, "Unsupported file type"

def extract_financial_data_with_patterns(document_text):
    """Extract financial data using pattern matching"""
    try:
        # Extract company name
        company_match = re.search(r'([A-Z][A-Za-z\s&]+(?:LIMITED|LTD|CORPORATION|CORP|INC|INDUSTRIES))', document_text)
        company_name = company_match.group(1).strip() if company_match else "Unknown"
        
        # Extract currency
        currency = "INR" if "₹" in document_text or "crores" in document_text.lower() else "USD"
        
        # Extract units
        units = "Crores" if "crores" in document_text.lower() else "Actual"
        
        # Find fiscal years - look for patterns like "2024", "2023", "9M 2025", "FY 2024"
        year_pattern = r'\b(20\d{2})\b'
        years = sorted(set(re.findall(year_pattern, document_text)), reverse=True)
        
        if not years:
            years = ["2024", "2023"]
        
        fiscal_years = years[:3]  # Take top 3 years
        
        financial_data = {}
        
        # Define line items to search for with various name variations
        line_items = {
            'revenue': [r'revenue\s+from\s+operations', r'total\s+revenue', r'sales', r'net\s+revenue', r'revenues'],
            'cost_of_revenue': [r'cost\s+of\s+(?:revenue|goods\s+sold|materials)', r'cogs', r'cost\s+of\s+sales'],
            'gross_profit': [r'gross\s+profit', r'gross\s+margin'],
            'operating_expenses': [r'operating\s+(?:expenses|costs)', r'sga', r'selling.*administrative'],
            'operating_income': [r'operating\s+(?:income|profit)', r'ebit'],
            'interest_expense': [r'interest\s+(?:expense|paid)'],
            'tax_expense': [r'(?:income\s+)?tax\s+(?:expense|cost)', r'provision\s+for\s+taxes'],
            'net_income': [r'net\s+(?:income|profit)', r'earnings', r'net\s+earnings'],
            'total_assets': [r'total\s+assets'],
            'total_liabilities': [r'total\s+liabilities'],
            'shareholders_equity': [r'(?:shareholders|stockholders)\s+equity', r'total\s+equity'],
        }
        
        # For each line item, search for values
        for item_key, patterns in line_items.items():
            financial_data[item_key] = {}
            
            for pattern in patterns:
                # Find the line containing this pattern
                lines = document_text.split('\n')
                for i, line in enumerate(lines):
                    if re.search(pattern, line, re.IGNORECASE):
                        # Extract numbers from this line and nearby lines
                        # Look for numbers in the current line
                        numbers = re.findall(r'[\d,]+\.?\d*', line)
                        
                        if numbers:
                            # Try to match with years
                            for j, year in enumerate(fiscal_years):
                                if j < len(numbers):
                                    try:
                                        value = float(numbers[j].replace(',', ''))
                                        financial_data[item_key][year] = value
                                    except:
                                        pass
                            break
        
        # Remove empty items
        financial_data = {k: v for k, v in financial_data.items() if v}
        
        result = {
            "company_name": company_name,
            "fiscal_years": fiscal_years,
            "financial_data": financial_data,
            "currency": currency,
            "units": units,
            "notes": [
                "Data extracted using pattern matching",
                "Some values may be approximate if document format varies",
                "Manual review recommended for accuracy"
            ]
        }
        
        return result, None
        
    except Exception as e:
        return None, f"Error extracting financial data: {str(e)}"

def create_excel_workbook(financial_data):
    """Create a professional Excel workbook from extracted financial data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Data"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    data_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    missing_font = Font(color="9C0006", italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"Financial Statement - {financial_data.get('company_name', 'Unknown Company')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Metadata
    ws['A2'] = f"Currency: {financial_data.get('currency', 'Unknown')}"
    ws['A3'] = f"Units: {financial_data.get('units', 'Actual')}"
    
    # Column headers
    years = financial_data.get('fiscal_years', [])
    ws['A5'] = "Line Item"
    ws['A5'].fill = header_fill
    ws['A5'].font = header_font
    ws['A5'].border = border
    
    for col_idx, year in enumerate(years, start=2):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = f"FY {year}"
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Financial items
    financial_items_list = [
        ('revenue', 'Revenue'),
        ('cost_of_revenue', 'Cost of Revenue'),
        ('gross_profit', 'Gross Profit'),
        ('operating_expenses', 'Operating Expenses'),
        ('operating_income', 'Operating Income'),
        ('interest_expense', 'Interest Expense'),
        ('tax_expense', 'Tax Expense'),
        ('net_income', 'Net Income'),
        ('total_assets', 'Total Assets'),
        ('total_liabilities', 'Total Liabilities'),
        ('shareholders_equity', 'Shareholders\' Equity'),
    ]
    
    data = financial_data.get('financial_data', {})
    
    row = 6
    for key, label in financial_items_list:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        
        for col_idx, year in enumerate(years, start=2):
            cell = ws.cell(row=row, column=col_idx)
            
            if key in data and year in data[key]:
                value = data[key][year]
                cell.value = value
                cell.number_format = '#,##0.00'
            else:
                cell.value = "N/A"
                cell.fill = missing_fill
                cell.font = missing_font
            
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Notes section
    row += 2
    ws[f'A{row}'] = "Notes & Assumptions:"
    ws[f'A{row}'].font = Font(bold=True)
    
    for note in financial_data.get('notes', []):
        row += 1
        ws[f'A{row}'] = f"• {note}"
        ws.merge_cells(f'A{row}:E{row}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    for col_idx in range(2, 2 + len(years)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    return wb

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/extract', methods=['POST'])
def extract():
    """Handle file upload and financial extraction"""
    
    # Check if file is present
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Only PDF, TXT, and DOCX files are supported'}), 400
    
    # Save uploaded file temporarily
    filename = secure_filename(file.filename)
    temp_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    try:
        file.save(temp_path)
        
        # Extract text from file
        document_text, error = extract_text_from_file(temp_path)
        if error:
            return jsonify({'error': error}), 400
        
        if not document_text or len(document_text.strip()) < 100:
            return jsonify({'error': 'Document appears to be empty or unreadable'}), 400
        
        # Extract financial data using pattern matching
        financial_data, error = extract_financial_data_with_patterns(document_text)
        if error:
            return jsonify({'error': error}), 400
        
        # Create Excel workbook
        workbook = create_excel_workbook(financial_data)
        
        # Save to bytes
        excel_bytes = BytesIO()
        workbook.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Return Excel file
        return send_file(
            excel_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"financial_extract_{financial_data.get('company_name', 'unknown').replace(' ', '_')}.xlsx"
        )
    
    finally:
        # Clean up temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)

@app.route('/api/health', methods=['GET'])
def health():
    """Health check for deployment"""
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)