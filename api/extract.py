from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import anthropic
import json
import os
import tempfile
from io import BytesIO
from pathlib import Path

ALLOWED_EXTENSIONS = {'pdf', 'txt', 'docx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_pdf(file_path):
    """Extract text from PDF using pdfplumber"""
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
    except Exception as e:
        return None, f"Error reading PDF: {str(e)}"
    return text, None

def extract_text_from_file(file_path):
    """Extract text from uploaded file"""
    file_ext = file_path.rsplit('.', 1)[1].lower()
    
    if file_ext == 'pdf':
        return extract_text_from_pdf(file_path)
    elif file_ext == 'txt':
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read(), None
        except Exception as e:
            return None, f"Error reading text file: {str(e)}"
    elif file_ext == 'docx':
        try:
            from docx import Document
            doc = Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
            return text, None
        except Exception as e:
            return None, f"Error reading DOCX: {str(e)}"
    
    return None, "Unsupported file type"

def extract_financial_data_with_claude(document_text):
    """Use Claude API to intelligently extract financial data"""
    try:
        client = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))
        
        prompt = f"""
You are a financial analyst. Extract financial statement data from this document.

Return ONLY a valid JSON object (no markdown, no code blocks) with this exact structure:
{{
  "company_name": "company name or 'Unknown'",
  "fiscal_years": ["2024", "2023", "2022"],
  "financial_data": {{
    "revenue": {{"2024": 123456.00, "2023": 120000.00}},
    "cost_of_revenue": {{"2024": 50000.00}},
    "gross_profit": {{"2024": 73456.00}},
    "operating_expenses": {{"2024": 30000.00}},
    "operating_income": {{"2024": 43456.00}},
    "interest_expense": {{}},
    "tax_expense": {{"2024": 8000.00}},
    "net_income": {{"2024": 35456.00}},
    "total_assets": {{"2024": 500000.00}},
    "total_liabilities": {{"2024": 200000.00}},
    "shareholders_equity": {{"2024": 300000.00}}
  }},
  "currency": "USD",
  "units": "thousands or actual",
  "notes": ["any missing data or assumptions"]
}}

IMPORTANT:
- Extract ONLY numbers that are explicitly stated in the document
- Do NOT hallucinate or estimate values
- If a line item is not found, omit it from the object
- Preserve the original numbers (don't convert units unless stated)
- If currency/units are unclear, note in "notes"
- Extract all years of data present
- Return ONLY the JSON object, nothing else

Document text:
{document_text[:8000]}
"""
        
        message = client.messages.create(
            model="claude-opus-4-5-20251101",
            max_tokens=2000,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        
        response_text = message.content[0].text.strip()
        financial_data = json.loads(response_text)
        return financial_data, None
        
    except json.JSONDecodeError as e:
        return None, f"Failed to parse financial data: {str(e)}"
    except Exception as e:
        return None, f"Claude API error: {str(e)}"

def create_excel_workbook(financial_data):
    """Create a professional Excel workbook from extracted financial data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Data"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    missing_font = Font(color="9C0006", italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"Financial Statement - {financial_data.get('company_name', 'Unknown Company')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Metadata
    ws['A2'] = f"Currency: {financial_data.get('currency', 'Unknown')}"
    ws['A3'] = f"Units: {financial_data.get('units', 'Actual')}"
    
    # Column headers
    years = financial_data.get('fiscal_years', [])
    ws['A5'] = "Line Item"
    ws['A5'].fill = header_fill
    ws['A5'].font = header_font
    ws['A5'].border = border
    
    for col_idx, year in enumerate(years, start=2):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = f"FY {year}"
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Financial items
    financial_items_list = [
        ('revenue', 'Revenue'),
        ('cost_of_revenue', 'Cost of Revenue'),
        ('gross_profit', 'Gross Profit'),
        ('operating_expenses', 'Operating Expenses'),
        ('operating_income', 'Operating Income'),
        ('interest_expense', 'Interest Expense'),
        ('tax_expense', 'Tax Expense'),
        ('net_income', 'Net Income'),
        ('total_assets', 'Total Assets'),
        ('total_liabilities', 'Total Liabilities'),
        ('shareholders_equity', 'Shareholders\' Equity'),
    ]
    
    data = financial_data.get('financial_data', {})
    
    row = 6
    for key, label in financial_items_list:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        
        for col_idx, year in enumerate(years, start=2):
            cell = ws.cell(row=row, column=col_idx)
            
            if key in data and year in data[key]:
                value = data[key][year]
                cell.value = value
                cell.number_format = '#,##0.00'
            else:
                cell.value = "N/A"
                cell.fill = missing_fill
                cell.font = missing_font
            
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Notes section
    row += 2
    ws[f'A{row}'] = "Notes & Assumptions:"
    ws[f'A{row}'].font = Font(bold=True)
    
    for note in financial_data.get('notes', []):
        row += 1
        ws[f'A{row}'] = f"• {note}"
        ws.merge_cells(f'A{row}:E{row}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    for col_idx in range(2, 2 + len(years)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    return wb

# Create Flask app for local development
app = Flask(__name__, template_folder='.', static_folder='.')
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024

@app.route('/')
def index():
    with open('templates/index.html', 'r') as f:
        return f.read()

@app.route('/api/extract', methods=['POST'])
def extract():
    """Handle file upload and financial extraction"""
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Only PDF, TXT, and DOCX files are supported'}), 400
    
    # Save uploaded file temporarily
    filename = secure_filename(file.filename)
    temp_path = os.path.join(tempfile.gettempdir(), filename)
    
    try:
        file.save(temp_path)
        
        # Extract text from file
        document_text, error = extract_text_from_file(temp_path)
        if error:
            return jsonify({'error': error}), 400
        
        if not document_text or len(document_text.strip()) < 100:
            return jsonify({'error': 'Document appears to be empty or unreadable'}), 400
        
        # Extract financial data using Claude
        financial_data, error = extract_financial_data_with_claude(document_text)
        if error:
            return jsonify({'error': error}), 400
        
        # Create Excel workbook
        workbook = create_excel_workbook(financial_data)
        
        # Save to bytes
        excel_bytes = BytesIO()
        workbook.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Return Excel file
        return send_file(
            excel_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"financial_extract_{financial_data.get('company_name', 'unknown').replace(' ', '_')}.xlsx"
        )
    
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    app.run(debug=True, port=5000)